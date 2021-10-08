# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import csv
import io
import re
import xlrd
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side, Alignment

import datetime


# Press the green button in the gutter to run the script.

workbook = load_workbook(u'C:/Users/taotao3\Desktop/01.xlsx')
booksheet = workbook.active

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns
i = 1
str='08:00:00'
date = time.strptime(str, "%H:%M:%S")
detester = '2019-05-01'
datetime1 = datetime.datetime.strptime(detester, '%Y-%m-%d')
todaymoney=53
tomomoney=0
todaymoney1=35
tomomoney1=20
# 迭代所有的行
for row in rows:
  i = i + 1
  line = [col.value for col in row]
  cell_data_1 = booksheet.cell(row=i, column=3).value        #  获取第i行1 列的数据
  cell_data_2 = booksheet.cell(row=i, column=4).value        # 获取第i行 2 列的数据
  cell_data_3 = booksheet.cell(row=i, column=5).value         # 获取第i行 3 列的数据

  if booksheet.cell(row=i, column=3).value == datetime1:
    if cell_data_2.hour < date.tm_hour:
      tomomoney=tomomoney+cell_data_3
    else:
      todaymoney= todaymoney +cell_data_3
  else:
    datetime1 = cell_data_1
    #booksheet.cell(row=todaymoney1, column=tomomoney1).value = 'test'
    booksheet.cell(row=todaymoney1-1, column=tomomoney1-1).value=datetime1
    booksheet.cell(row=todaymoney1, column=tomomoney1).value=todaymoney
    #print(todaymoney1, tomomoney1)
    todaymoney1 = todaymoney1-1
   # if i != 2:
    todaymoney = 0
    if datetime1!=None:
      todaymoney=todaymoney+tomomoney+cell_data_3
    tomomoney = 0
workbook.save('C:/Users/taotao3/Desktop/01.xlsx')
  # print(cell_data_1, cell_data_2, cell_data_3)